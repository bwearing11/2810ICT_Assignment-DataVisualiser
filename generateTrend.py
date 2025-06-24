#%%
import openpyxl
import matplotlib.pyplot as plt
from collections import Counter
import numpy


wb = openpyxl.load_workbook('C:/Users/benja/.spyder-py3/penalty_data_set_2.xlsx')
ws = wb.active
                                                                          #Generate a trendline for a particular offence code
def generateTrend():                                        
    filt = ('Stop within 10 metres of an intersection without traffic lights',)     
    lst = []
    years = ['2011-2012', '2012-2013', '2013-2014', '2014-2015', '2015-2016', '2016-2017', '2017-2018']
    
    i = 0
    for row in ws.iter_rows(min_col = 4, max_col = 4, values_only = True):      #filters the data 
        i+= 1
        if row == filt:
            j = str(i)
            dateVal = 'A'+j
            date = ws[dateVal].value
            lst.append(date)  
                                                                                
    count = Counter(lst)
                                                    #each variable is a count of no of offences                                                                                 #for the filter. No of offences per year                             
    x0 = count['2011-2012']                                                     #is put into these counts
    x1 = count['2012-2013']
    x2 = count['2013-2014']
    x3 = count['2014-2015']
    x4 = count['2015-2016']
    x5 = count['2016-2017']
    x6 = count['2017-2018']
    
    vals = [x0, x1, x2, x3, x4, x5, x6]
    
    
    plt.bar(years, vals)
    plt.title(filt)
    plt.show() 

          
        
generateTrend()
    
def generateTable():
    filt = ('Stop within 10 metres of an intersection without traffic lights',)     
    lst = []
    years = ['2011-2012', '2012-2013', '2013-2014', '2014-2015', '2015-2016', '2016-2017', '2017-2018']
    
    i = 0
    for row in ws.iter_rows(min_col = 4, max_col = 4, values_only = True):      #filters the data 
        i+= 1
        if row == filt:
            j = str(i)
            dateVal = 'A'+j
            date = ws[dateVal].value
            lst.append(date)  

#lst.count()
# plot offences over time
# display to screen
#plt.show()

